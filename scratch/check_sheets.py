import openpyxl

def inspect_file(path):
    print(f"\nInspecting {path}...")
    wb = openpyxl.load_workbook(path, read_only=True)
    print(f"Sheet names: {wb.sheetnames}")
    for name in wb.sheetnames[:3]:
        sheet = wb[name]
        print(f"\nSheet: {name}")
        # Print first 5 rows
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if idx > 5:
                break
            print(f"  Row {idx}: {row}")

inspect_file(r"C:\Users\에스에스브이\.gemini\antigravity\scratch\excel-exporter\sample\sample.xlsx")
inspect_file(r"C:\Users\에스에스브이\.gemini\antigravity\scratch\excel-exporter\sample\sample2.xlsx")
