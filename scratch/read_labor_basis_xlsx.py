import openpyxl
import json

path = r"C:\Users\에스에스브이\.gemini\antigravity\scratch\excel-exporter\sample\sample2.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

# Find the sheet that corresponds to "노임근거"
sheet_name = None
for name in wb.sheetnames:
    # "노임근거" matches in Korean
    if "노임" in name or "근거" in name:
        sheet_name = name
        break

if not sheet_name:
    print("No labor basis sheet found!")
    exit(1)

sheet = wb[sheet_name]
print(f"Reading sheet: {sheet_name}")

rows_data = []
for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
    rows_data.append(row)

# Print first 20 rows of data
for idx, r in enumerate(rows_data[:25]):
    print(f"Row {idx+1}: {r}")

# Let's save it to a JSON file to inspect
out_json = r"C:\Users\에스에스브이\.gemini\antigravity\scratch\excel-exporter\scratch\excel_labor_basis.json"
with open(out_json, "w", encoding="utf-8") as f:
    json.dump(rows_data, f, ensure_ascii=False, indent=4)

print(f"\nSaved all rows to {out_json}")
